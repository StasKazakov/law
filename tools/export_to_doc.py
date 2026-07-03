import asyncio
import pandas as pd
from utils.db_connection import init_db, close_db
import openpyxl
from openpyxl.styles import Alignment, Font

async def export_benchmark_to_excel():
    """Fetch successfully generated rows with cause_num and export to styled Excel file"""
    pool = await init_db()
    
    # ORDER OF COLUMNS MATCHES THE PANDAS DATAFRAME COLUMNS EXACTLY
    query = """
        SELECT doc_id, cause_num, court_code, justice_kind, question, doc_url
        FROM doc_eval_100
        WHERE question IS NOT NULL
        ORDER BY id;
    """
    
    print("Fetching data from the database...")
    async with pool.acquire() as conn:
        rows = await conn.fetch(query)
        
    if not rows:
        print("[WARNING] No generated questions found in the database yet.")
        print("[TIP] If you want to test the script before running Gemini, temporarily remove 'WHERE question IS NOT NULL'.")
        await close_db()
        return

    # Convert records to dictionary format
    data = [dict(row) for row in rows]
    df = pd.DataFrame(data)
    
    # Setting precise column names mapping 1:1 with SELECT statement
    df.columns = [
        "Document ID", 
        "Case Number (Cause Num)", 
        "Court Code", 
        "Justice Kind", 
        "Generated Question", 
        "Document URL"
    ]
    
    output_filename = "legal_benchmark_examples.xlsx"
    print(f"Exporting {len(df)} records to {output_filename}...")
    
    # Basic export using openpyxl engine
    df.to_excel(output_filename, index=False, engine='openpyxl')
    
    # --- Advanced formatting for text wrapping and link styling ---
    try:
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active
        
        # Enable text wrap for Questions and URLs to make it human-readable
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            # Question cell wrapping
            row[4].alignment = Alignment(wrap_text=True, vertical="top")
            # URL cell styling (blue, underlined, wrapped)
            row[5].alignment = Alignment(wrap_text=True, vertical="top")
            row[5].font = Font(color="0563C1", underline="single")
            
        # Set dynamic column widths so the fields look clean
        ws.column_dimensions['A'].width = 15  # Document ID
        ws.column_dimensions['B'].width = 25  # Case Number
        ws.column_dimensions['C'].width = 15  # Court Code
        ws.column_dimensions['D'].width = 15  # Justice Kind
        ws.column_dimensions['E'].width = 65  # Generated Question (wide for long text)
        ws.column_dimensions['F'].width = 40  # Document URL
        
        wb.save(output_filename)
    except Exception as e:
        print(f"[NOTE] Excel layout styling skipped, but data saved: {e}")

    print("[SUCCESS] Excel file generated successfully!")
    await close_db()

if __name__ == "__main__":
    asyncio.run(export_benchmark_to_excel())